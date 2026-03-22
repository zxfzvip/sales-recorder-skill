import os
from openpyxl import load_workbook
import re

def record_inventory(target: str, date: str = None, product: str = None, qty: float = None, price: float = None, expr_count: float = None, expr_price: float = None):
    base_path = "/Users/mac/Desktop/拿货记数"
    file_path = os.path.join(base_path, f"{target}.xlsx")
    if not os.path.exists(file_path):
        return f"找不到文件：{target}.xlsx"
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        next_row = 2
        for row in range(2, 1000):
            val_b = ws.cell(row, 2).value
            val_f = ws.cell(row, 6).value
            is_b_empty = val_b is None or (isinstance(val_b, str) and val_b.startswith('='))
            is_f_empty = val_f is None or (isinstance(val_f, str) and val_f.startswith('='))
            if is_b_empty and is_f_empty:
                next_row = row
                break
        
        # 日期填入A列
        if date:
            ws.cell(next_row, 1).value = date
        
        if product:
            ws.cell(next_row, 2).value = product
            ws.cell(next_row, 3).value = qty or 0
            ws.cell(next_row, 4).value = price or 0
            ws.cell(next_row, 5).value = f"=C{next_row}*D{next_row}"
        if expr_count is not None:
            ws.cell(next_row, 6).value = expr_count
            ws.cell(next_row, 7).value = expr_price or 0
            ws.cell(next_row, 8).value = f"=F{next_row}*G{next_row}"
        
        wb.save(file_path)
        
        # 计算小计
        subtotal = (qty or 0) * (price or 0) if qty and price else None
        expr_subtotal = (expr_count or 0) * (expr_price or 0) if expr_count and expr_price else None
        
        # 构建回复 - 表格形式
        result = f"✅ 已添加到 {target}.xlsx 第{next_row}行！\n\n"
        
        # 表头
        result += "| 行 | 日期 | 商品 | 数量 | 单价 | 小计 | 快递 | 快递费 | 快递小计 |\n"
        result += "|----|------|------|------|------|------|------|--------|----------|\n"
        
        # 商品行
        if product and qty and price:
            p_row = f"| {next_row} | {date or ''} | {product} | {int(qty)} | {price} | ¥{subtotal:.0f} |"
        else:
            p_row = f"| {next_row} | {date or ''} | | | | |"
        
        # 快递行
        if expr_count and expr_price:
            e_row = f"| | | | | | | {int(expr_count)} | {expr_price} | ¥{expr_subtotal:.0f} |"
        else:
            e_row = ""
        
        result += p_row + "\n"
        if e_row:
            result += e_row + "\n"
        
        result += "\n继续 📝"
        return result
        
    except Exception as e:
        return f"错误: {str(e)}"

def parse_args(msg: str):
    """解析用户消息，提取参数"""
    # 提取目标表格
    tables = ["弟弟", "央央", "宝宝", "超宝", "薛泽凯", "锐", "凯"]
    target = "弟弟"  # 默认
    for t in tables:
        if t in msg:
            target = t
            break
    
    # 提取日期：匹配 X月Y号 或 Y号 格式（必须带"号"才是日期）
    date_match = re.search(r'(\d+)月(\d+)号|(\d+)号', msg)
    date = None
    if date_match:
        if date_match.group(1):  # X月Y号格式
            month = date_match.group(1)
            day = date_match.group(2)
            date = f"{month}月{day}号"
        else:  # Y号格式
            day = date_match.group(3)
            date = f"{day}号"
    
    # 提取商品
    products = ["风流果", "风流", "润滑油", "润滑", "润滑液", "高潮液", "延时喷剂", "延时膏", "阳具", "川井 依克多因", "川井"]
    product = None
    for p in products:
        if p in msg:
            product = p
            break
    
    # 如果没匹配到常见商品，看是否有"数量"前面的文字
    if not product:
        qty_match = re.search(r'(\S+)\s*数量', msg)
        if qty_match:
            potential = qty_match.group(1)
            # 排除日期和表格名
            if potential not in ["3", "4", "5", "6", "7", "8", "9", "10"] + tables:
                product = potential
    
    # 提取数量
    qty_match = re.search(r'数量(\d+)', msg)
    qty = float(qty_match.group(1)) if qty_match else None
    
    # 提取快递数量和快递价格（先提取快递相关）
    expr_match = re.search(r'快递(\d+)价格(\d+\.?\d*)', msg)
    if expr_match:
        expr_count = float(expr_match.group(1))
        expr_price = float(expr_match.group(2))
    else:
        # 只匹配快递数量
        expr_count_match = re.search(r'快递(\d+)', msg)
        expr_count = float(expr_count_match.group(1)) if expr_count_match else None
        expr_price = None
    
    # 提取商品价格
    # 优先提取"商品价格"中的价格
    if product:
        price_match = re.search(rf'{re.escape(product)}.*?价格(\d+\.?\d*)', msg)
        if not price_match:
            # 尝试从"数量X价格Y"中提取
            price_match = re.search(r'数量\d+价格(\d+\.?\d*)', msg)
        price = float(price_match.group(1)) if price_match else None
    else:
        # 没有商品时，价格应该归快递
        price = None
    
    return target, date, product, qty, price, expr_count, expr_price

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 2:
        # 命令行参数模式
        target = sys.argv[1] if len(sys.argv) > 1 else "弟弟"
        product = sys.argv[2] if len(sys.argv) > 2 else None
        qty = float(sys.argv[3]) if len(sys.argv) > 3 and sys.argv[3] else None
        price = float(sys.argv[4]) if len(sys.argv) > 4 and sys.argv[4] else None
        expr_count = float(sys.argv[5]) if len(sys.argv) > 5 and sys.argv[5] else None
        expr_price = float(sys.argv[6]) if len(sys.argv) > 6 and sys.argv[6] else None
        print(record_inventory(target, None, product, qty, price, expr_count, expr_price))
    else:
        # 消息解析模式
        msg = " ".join(sys.argv[1:]) if len(sys.argv) > 1 else ""
        target, date, product, qty, price, expr_count, expr_price = parse_args(msg)
        
        if not product and not expr_count:
            print("无法识别商品或快递，请重试")
        else:
            print(record_inventory(target, date, product, qty, price, expr_count, expr_price))
